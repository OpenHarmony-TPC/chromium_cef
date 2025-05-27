import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from author_info import author_group_transform_preset
import pandas as pd
import io
import subprocess
import uuid

'''
# 脚本功能：
1. 依据git log所导出的日志，统计日志中每次commit侵入式修改的文件及修改行数
2. 侵入式修改即src/arkweb、src/cef/ohos_cef_ext、deps_code、src/third_party/rust-toolchain等目录以外的修改

# 使用步骤：
1. 安装openpyxl 安装pandas
    > pip3 install pandas openpyxl -i https://mirrors.huaweicloud.com/repository/pypi/simple
2. terminal 切换到 chromium_132.0.6789.0_single 本地目录内，**手动**更新本地代码至最新代码
3. 运行程序
    > python3 /path/to/your/chromium_132_log_parse.py
4. 程序输出
    > 在terminal当前目录生成output_2025-01-22T11-45-39.885247.xlsx格式excel文件

# 补充信息：
一、程序内所用到的git获取日志文件的指令
1. git log --oneline，能比较方便的提取commit id与提交信息的对应关系
    > git log --oneline | less -S > chromium_132_git_log_one_line.log;
2. git log --stat，能输出每次commit中对每个文件的修改行数
    > git log --stat --stat-width=1000  | less -S > chromium_132_git_log_stat.log
二、author_info.py
1. git log --stat中的“Author”、“姓名 工号”、“所属团队”三者的对应信息取自本文件同目录下author_info.py，需手动实时更新
2. 若所导出excel表格责任人一列显示不是“姓名 工号”，表示对应的责任人信息未更新到author_info.py，需手动更新
3. excel第二页“责任人及团队”仅统计第一页提交了”侵入式修改“的责任人，不是全部责任人
'''


# git log --oneline选项日志
def get_git_log_one_line_data():
    random_file_name = 'temporary_git_log_one_line_' + datetime.datetime.now().isoformat().replace(':', '-') + '.log'
    command = f'git log --oneline | less -S > {random_file_name}'
    print(f'\t正在创建临时文件：{os.path.abspath(random_file_name)}')
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    print(f'\t临时文件已创建')
    with open(random_file_name, "r", encoding='utf-8') as f:
        commit_info_list = f.readlines()
    command = f'rm {random_file_name}'
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    print(f'\t临时文件已删除')
    commit_info_list = [i.strip() for i in commit_info_list]
    commit_info_dict = {}
    for i in commit_info_list:
        commit_id = i.split(' ')[0]
        commit_info = ' '.join(i.split(' ')[1:])
        commit_info_dict[commit_id] = commit_info
    assert len(commit_info_list) == len(set(commit_info_dict.keys())), 'commit id有重复'
    return commit_info_dict

# git log --stat选项日志
def get_git_log_stat_data():
    random_file_name = 'temporary_git_log_stat_' + datetime.datetime.now().isoformat().replace(':', '-') + '.log'
    command = f'git log --stat --stat-width=1000  | less -S > {random_file_name}'
    print(f'\t正在创建临时文件：{os.path.abspath(random_file_name)}')
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    print(f'\t临时文件已创建')
    with open(random_file_name, 'r', encoding='utf-8') as f:
        data_stat_list = f.readlines()
    command = f'rm {random_file_name}'
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    print(f'\t临时文件已删除')
    data_stat_list = [i.strip() for i in data_stat_list]
    return data_stat_list

def log_parse(
        commit_id_list, commit_id_hyperlink_list, file_name_list, edited_lines_list, commit_info_list, author_list_git, author_list_transformed,
        commit_info_dict, data_stat_list,
        target_file_suffix, removed_commid_id_list
):
    line_index = 0
    percent_process = 0
    def process_percent_record(line_index, percent_process):
        if line_index * 10 // len(data_stat_list) > percent_process:
            percent_process = line_index * 10 // len(data_stat_list)
            print(f'\t日志解析进度：{percent_process*10:>2}%')
        return percent_process
    while line_index < len(data_stat_list):
        percent_process = process_percent_record(line_index, percent_process)
        while line_index < len(data_stat_list) and not data_stat_list[line_index].startswith('commit '):
            line_index += 1
            percent_process = process_percent_record(line_index, percent_process)
        if line_index >= len(data_stat_list):
            break
        if data_stat_list[line_index + 1].startswith('Merge: '):
            line_index += 1
            continue
        commit_id = data_stat_list[line_index][7:]
        if commit_id in removed_commid_id_list:
            print(f'\t跳过：{commit_id}')
            line_index += 1
            continue
        count = 0
        commit_info = ''
        for i in commit_info_dict:
            if commit_id.startswith(i):
                commit_info = commit_info_dict[i]
                count += 1
        assert count == 1, 'git log --one-line与git log --stat的commit id数目不一致'

        line_index += 1
        author_git = data_stat_list[line_index][8:]
        line_index += 1

        author_transformed = author_git
        count = 0
        for i in author_group_transform_preset:
            if author_git == i[0]:
                author_transformed = i[1]
                count += 1
        assert count <= 1, f'author_info.py中author信息有重复请手动去重, 重复责任人信息为{author_git}'

        line_index += 1
        while '|' not in data_stat_list[line_index]:
            line_index += 1
            percent_process = process_percent_record(line_index, percent_process)
        while '|' in data_stat_list[line_index]:
            item = data_stat_list[line_index].split('|')
            file_name = item[0].strip()

            if os.path.splitext(item[0])[1].strip() not in target_file_suffix:
                line_index += 1
                continue

            # 排除不属于侵入式修改的目录
            if file_name.startswith('src/arkweb/') or file_name.startswith('src/cef/ohos_cef_ext/') \
                    or file_name.startswith('deps_code/') or file_name.startswith('src/third_party/rust-toolchain/') \
                    or file_name.startswith('src/build/rust/std/rules/') \
                    or file_name.startswith('src/cef/libcef_dll/') or file_name.startswith('src/ohos/') \
                    or file_name.startswith('src/{arkweb/chromium_ext/chrome => cef/ohos_cef_ext') \
                    or file_name.startswith('src/{ => arkweb}') or file_name.startswith('src/{ => arkweb/') \
                    or file_name.startswith('src/{services => arkweb/') \
                    or file_name.startswith('src/{components/viz/service => arkweb/') \
                    or file_name.startswith('src/{ohos => arkweb/') \
                    or file_name.startswith('src/cef/{ => ohos_cef_ext/') \
                    or file_name.startswith('src/cef/{ => ohos_cef_ext}'):
                line_index += 1
                continue

            # 二进制文件，跳过
            if item[1].strip().startswith('Bin'):
                line_index += 1
                continue
            commit_id_list.append(commit_id)
            commit_id_hyperlink_list.append(
                r'=HYPERLINK("https://open.codehub.huawei.com/innersource/shanhai/wutong/chromium/files/commit/' + commit_id + r'?ref=chromium_132.0.6789.0_single","' + commit_id + '")')
            file_name_list.append(file_name)
            edit_lines = eval(item[1].strip().split(' ')[0])
            edited_lines_list.append(edit_lines)
            commit_info_list.append(commit_info)
            author_list_git.append(author_git)
            author_list_transformed.append(author_transformed)

            line_index += 1
            percent_process = process_percent_record(line_index, percent_process)

        if line_index + 2 < len(data_stat_list):
            assert data_stat_list[line_index + 2].startswith('commit '), '提交信息包含"|"，当前解析日志逻辑不能正确处理，请手动检查或修改代码'

def author_group_info_export(author_list_transformed):
    author_group_transform_dict = {}
    for i in author_group_transform_preset:
        author_group_transform_dict[i[1]] = i[2]
    author_group_export = []
    for i in author_list_transformed:
        if i in author_group_transform_dict.keys():
            author_group_export.append((i, author_group_transform_dict[i]))
        else:
            author_group_export.append((i, ''))
    author_group_export = list(set(author_group_export))
    author_group_export.sort(key=lambda x: x[0])
    author_group_export.sort(key=lambda x: x[1])

    excel_author_list = []
    excel_group_list = []
    for i in author_group_export:
        excel_author_list.append(i[0])
        excel_group_list.append(i[1])

    return excel_author_list, excel_group_list

def result_save(commit_id_hyperlink_list, file_name_list, edited_lines_list, commit_info_list, author_list_transformed, excel_author_list, excel_group_list, ):
    df = pd.DataFrame({
        '提交id': commit_id_hyperlink_list,
        '文件路径': file_name_list,
        '修改行数': edited_lines_list,
        '提交信息': commit_info_list,
        '责任人': author_list_transformed
    })
    df_author_group = pd.DataFrame({
        '责任人': excel_author_list,
        '所属团队': excel_group_list
    })
    io_file = io.BytesIO()
    with pd.ExcelWriter(io_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='提交日志', index=False)
        df_author_group.to_excel(writer, sheet_name='责任人及团队', index=False)

    wb = load_workbook(io_file)
    ws = wb['提交日志']
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(color='0000FF', underline='single')
    file_save_name = f'output_{datetime.datetime.now().isoformat().replace(":", "-")}.xlsx'
    wb.save(file_save_name)
    return os.path.abspath(file_save_name)

def print_commit_files(commit_id_list):
    commit_id_count_dict = {}
    for i in commit_id_list:
        if i in commit_id_count_dict:
            commit_id_count_dict[i] += 1
        else:
            commit_id_count_dict[i] = 1
    commit_id_count_list = list(commit_id_count_dict.items())
    commit_id_count_list.sort(key=lambda x: x[1], reverse=True)
    for i in commit_id_count_list:
        print(f'{i[0]}\t{i[1]}')

def print_author_lines(author_list_transformed):
    author_count_dict = {}
    for i in author_list_transformed:
        if i in author_count_dict:
            author_count_dict[i] += 1
        else:
            author_count_dict[i] = 1
    author_count_list = list(author_count_dict.items())
    author_count_list.sort(key=lambda x: x[1], reverse=True)
    for i in author_count_list:
        print(f'{i[0]:<30}\t{i[1]}')

if __name__ == '__main__':
    # 所筛选文件类型白名单
    target_file_suffix = [
        '.h', '.cc', '.cpp', '.c', '.hpp', '.mojom', '.gni', '.gn',
        '.py', '.java', '.go', '.xml', '.test', '.sh', '.cmake', '.headers',
        '.cfg', '.php', '.yml', '.def', '.patch', '.tmpl', '.filter', '.swift',
        '.jsx', '.flags', '.pyl', '.filelist', '.gradle', '.include', '.cnf',
        '.options', '.manifest', '.gypi', '.bash', '.list', '.bnf', '.config',
        '.targets', '.defs', '.fst', '.swf', '.gyp', '.task', '.gen', '.bats', '.msc',
        '.make', '.batch', '.py3', '.dist', '.hap', '.cef', '.allowlist', '.gin', '.args',
        '.opts', '.cmakein', '.exclude', '.protojson', '.unittests', '.vconf', '.c++', '.core'
    ]
    # 所排除commit id（把单次commit中文件修改数量超过2w个的commit记录排除）
    removed_commid_id_list = [
        '584d756a6c9779916f0c508c60c5b580f1225b7a',  # [Huawei] repo init Offering:oh
        'cd6492bea3f21627fdc8fcf8630be288fd275d23',  # Revert "upgrade chromium from 132.0.6789.0 to 132.0.6834.89"
        'fb25955a73ff25dbe593642a32b7bf2118be0ec0',  # upgrade chromium from 132.0.6789.0 to 132.0.6834.89
        'e94af7ffd4f4f594eb5ad4a978e4fe07503d7a16',  # [Huawei] repo init: upload gitignore files Offering:oh
    ]
    commit_id_list = []
    commit_id_hyperlink_list = []
    file_name_list = []
    edited_lines_list = []
    commit_info_list = []
    author_list_git = []
    author_list_transformed = []

    command = f'git log -1'
    result = subprocess.run(command, shell=True, text=True, capture_output=True)
    assert not result.stderr, f'请在git目录下执行本文件，result.stderr：{result.stderr}'


    print(f'\033[34m**白名单文件类型：\033[0m{target_file_suffix}\n')
    print(f'\033[34m**所排除commit id：\033[0m{removed_commid_id_list}\n')

    print(f'\033[34m1. 开始提取git log --oneline | less -S\033[0m')
    s_t = datetime.datetime.now()
    commit_info_dict = get_git_log_one_line_data()
    e_t = datetime.datetime.now()
    print(f'\t已获取git log --oneline数据，耗时 {e_t - s_t}')


    print(f'\033[34m2. 开始提取git log --stat --stat-width=1000  | less -S\033[0m')
    s_t = datetime.datetime.now()
    data_stat_list = get_git_log_stat_data()
    e_t = datetime.datetime.now()
    print(f'\t已获取git log --stat数据，耗时 {e_t - s_t}')


    print(f'\033[34m3. 开始解析日志\033[0m')
    s_t = datetime.datetime.now()
    log_parse(commit_id_list, commit_id_hyperlink_list, file_name_list, edited_lines_list, commit_info_list, author_list_git, author_list_transformed, commit_info_dict, data_stat_list, target_file_suffix, removed_commid_id_list)
    e_t = datetime.datetime.now()
    print(f'\t已解析完成，耗时 {e_t - s_t}')


    print(f'\033[34m4. 开始提取存在侵入式修改的责任人与所属团队信息\033[0m')
    s_t = datetime.datetime.now()
    excel_author_list, excel_group_list = author_group_info_export(author_list_transformed)
    e_t = datetime.datetime.now()
    print(f'\t已提取，耗时 {e_t - s_t}')


    print(f'\033[34m5. 开始导出excel表格\033[0m')
    s_t = datetime.datetime.now()
    excel_path = result_save(commit_id_hyperlink_list, file_name_list, edited_lines_list, commit_info_list, author_list_transformed, excel_author_list, excel_group_list)
    e_t = datetime.datetime.now()
    print(f'\t\033[32m已导出：{excel_path}\033[0m，耗时 {e_t - s_t}')

    # print_commit_files(commit_id_list)
    # print_author_lines(author_list_transformed)