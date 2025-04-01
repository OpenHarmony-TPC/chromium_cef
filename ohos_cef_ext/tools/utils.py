import json
from openpyxl.styles import Font, colors
from openpyxl import Workbook

def ReadConfigure():
    try:
        # 读取数据
        with open('src/cef/ohos_cef_ext/tools/compare.config.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print('配置文件不存在')
    return {}


def SafeGetVal(obj, key):
    try:
        return obj[key]
    except KeyError as e:
        return ''

class ReportTbl:
    def __init__(self):
        self.wb = Workbook()
    
    def addFirstSheet(self, data, title):
        # 创建第一个sheet
        sheet1 = self.wb.active
        sheet1.title = title
        for row in data:
            sheet1.append(row)
        
        cell0 = sheet1.cell(2, 2)
        # 设置字体对象 颜色为红色
        font_ins = Font(name="微软雅黑", sz=18, family=2, b=True, i=True, color=colors.Color(index=2))
        # 将字体对象赋予单元格
        cell0.font = font_ins

        for x in range(1,5):
            cell0 = sheet1.cell(len(data), x)
            # 设置字体对象 颜色为红色
            font_ins = Font(name="微软雅黑", sz=11, family=2, b=True)
            # 将字体对象赋予单元格
            cell0.font = font_ins


    def addSheet(self, data, title):
        # 新增sheet
        ws = self.wb.create_sheet(title)
        for row in data:
            ws.append(row)

        for x in range(1,10):
            cell0 = ws.cell(1, x)
            # 设置字体对象 颜色为红色
            font_ins = Font(name="微软雅黑", sz=11, family=2, b=True, i=True)
            # 将字体对象赋予单元格
            cell0.font = font_ins

    def save(self, filename):
        self.wb.save(filename)


def ReadAuthorsConfigure():
    try:
        # 读取数据
        with open('src/cef/ohos_cef_ext/tools/authors.config.json', 'r') as f:
            data = json.load(f) 
            return data['authArr'], data['mails']
    except FileNotFoundError:
        print('配置文件不存在')
    return [],[]

authArr,mailArr = ReadAuthorsConfigure()

def FindGroupbyID(authid): # wx1158027 l30063680
    if 'WX' not in authid:
        authid = authid[1:]

    for item in authArr:
        if authid in item[0]:
            return item

    return authid, '', ''

def FindGroupbyEMail(email): # wx1158027 l30063680
    for item in mailArr:
        if email in item[0]:
            return FindGroupbyID(item[1])

    return email, '', ''